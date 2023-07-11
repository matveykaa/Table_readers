from abc import ABC, abstractmethod
import csv
import openpyxl

CellValue = str | int | float

class TableReader(ABC):
    @abstractmethod
    def __init__(self, filename: str):
        pass

    @abstractmethod
    def read(self) -> list[list[CellValue]]:
        pass

class CSVReader(TableReader):
    def __init__(self, filename: str):
        self._filename = filename

    def read(self) -> list[list[CellValue]]:
        info_list: list[list[CellValue]] = []
        with open(self._filename, 'r') as fd:
            file_reader = csv.reader(fd, delimiter=',')
            for row in file_reader:
                info_list.append(row)
        print(info_list)
        return info_list

class XLSXReader(TableReader):
    def __init__(self, filename: str):
        self._filename = filename
    def read(self) -> list[list[CellValue]]:
        info_list: list[list[CellValue]] = []
        wb = openpyxl.load_workbook(self._filename)
        ws = wb.active
        last_row = ws.max_row
        last_column = ws.max_column
        for i in range(1, last_row+1):
            row_data = []
            for j in range(1, last_column+1):
                row_data.append(ws.cell(row=i, column=j).value)
            info_list.append(row_data)
        print(info_list)
        return info_list


if __name__ == "__main__":

    reader = CSVReader('table_csv.csv')
    reader.read()

    reader = XLSXReader('table.xlsx')
    reader.read()
